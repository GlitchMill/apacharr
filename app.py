import tempfile
import os
from flask import Flask, render_template, request, redirect, url_for, flash, Response, session, send_file
import openpyxl
from fpdf import FPDF
from io import BytesIO
import random
from tempfile import NamedTemporaryFile

app = Flask(__name__)
app.secret_key = os.getenv('SECRET_KEY')  # Secret key for session management

# Temporary directory for storing uploaded files
TEMP_UPLOAD_FOLDER = tempfile.mkdtemp()

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() == 'xlsx'

def check_excel_format(file):
    try:
        workbook = openpyxl.load_workbook(file)
        sheet = workbook.active
        headers = [str(cell.value).strip().lower() for cell in sheet[1] if cell.value is not None and cell.value.strip()]

        expected_headers = ['unit', 'questions', 'marks', 'type of question', 'probability']
        return headers == expected_headers
    except Exception as e:
        print(f"Error: {e}")
        return False

def get_question_types(file):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook.active
    question_types = {}

    for row in sheet.iter_rows(min_row=2, values_only=True):
        q_type = row[3]
        if q_type is not None:
            question_types[q_type] = question_types.get(q_type, 0) + 1

    return question_types

def generate_question_paper(file, request_data):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook.active

    questions = {row[3]: [] for row in sheet.iter_rows(min_row=2, values_only=True) if row[3] is not None}
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if row[3] is not None:
            questions[row[3]].append(row)

    selected_questions = []
    for q_type, num_questions in request_data.items():
        num_questions = int(num_questions)
        if num_questions > 0 and q_type in questions:
            selected_questions.extend(random.sample(questions[q_type], min(num_questions, len(questions[q_type]))))

    return selected_questions

def create_pdf(questions):
    pdf = FPDF()
    pdf.set_auto_page_break(auto=True, margin=15)
    pdf.add_page()
    pdf.set_font("Arial", size=12)

    # Group questions by type
    grouped_questions = {}
    for row in questions:
        unit, question, marks, q_type, probability = row[:5]
        if q_type not in grouped_questions:
            grouped_questions[q_type] = []
        grouped_questions[q_type].append((unit, question, marks))

    for q_type, q_list in grouped_questions.items():
        total_marks = sum(int(marks) for _, _, marks in q_list)  # Calculate total marks
        num_questions = len(q_list)

        # Add section header
        pdf.cell(0, 10, txt=f"{q_type} - {total_marks} x {num_questions} = {total_marks}", ln=True)

        # Add each question
        for index, (unit, question, marks) in enumerate(q_list, start=1):
            pdf.cell(0, 10, txt=f"{index}. {question} ({marks} marks)", ln=True)

        pdf.cell(0, 10, txt="", ln=True)  # Add a blank line between question types

    # Create a temporary file to hold the PDF data
    with tempfile.NamedTemporaryFile(delete=False, suffix='.pdf') as temp_pdf:
        pdf.output(temp_pdf.name)  # Save PDF to temporary file
        temp_pdf.seek(0)  # Move the cursor back to the beginning of the file

        # Read the content into a BytesIO object
        pdf_output = BytesIO(temp_pdf.read())

    return pdf_output

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/upload', methods=['POST'])
def upload_file():
    if 'file' not in request.files:
        flash('No file part')
        return redirect(request.url)

    file = request.files['file']

    if file.filename == '':
        flash('No selected file')
        return redirect(request.url)

    if file and allowed_file(file.filename):
        # Save the file temporarily to the disk
        temp_filepath = os.path.join(TEMP_UPLOAD_FOLDER, file.filename)
        file.save(temp_filepath)

        if check_excel_format(temp_filepath):
            # Store the file path in the session (instead of the file itself)
            session['file_path'] = temp_filepath
            question_types = get_question_types(temp_filepath)
            return render_template('question_selection.html', question_types=question_types)
        else:
            flash('File uploaded but does not match the expected format.')
            return redirect(url_for('index'))

    flash('Invalid file type. Please upload an .xlsx file.')
    return redirect(request.url)

@app.route('/generate', methods=['POST'])
def generate():
    selected_questions = []
    
    # Get the file path from the session
    file_path = session.get('file_path')
    
    if not file_path:
        flash('No file found in session.')
        return redirect(url_for('index'))

    for q_type in request.form:
        if q_type.endswith('_count'):
            count = int(request.form[q_type])  # Extract the number of questions
            question_type = q_type[:-6]  # Remove '_count' from the end
            selected_questions.extend(generate_question_paper(file_path, {question_type: count}))

    if not selected_questions:
        flash('No questions selected. Please try again.')
        return redirect(url_for('index'))

    # Create PDF with the selected questions
    pdf_output = create_pdf(selected_questions)

    # Return the PDF file as an attachment
    return send_file(pdf_output, mimetype='application/pdf', as_attachment=True, download_name='question_paper.pdf')


if __name__ == '__main__':
    app.run(debug=True)
